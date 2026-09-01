import csv
import hashlib
import logging
import re
import secrets
import string
from collections import Counter
from datetime import datetime, timedelta, timezone
from email.utils import format_datetime
from io import BytesIO, StringIO
from typing import Any, Callable, Iterable, Iterator

from app.auth.utils import hash_password
from app.settings.utils import get_value_by_page_and_key
from app.users.init import update_user_settings_bulk
from app.users.models import (
    create_user,
    user_exists_by_email,
)
from app.users.roles import (
    USER_ROLE,
)
from app.users.schemas import UserCreate
from app.users.utils import (
    _assert_password_policy,
    get_password_policy_requirements,
)
from fastapi import HTTPException, Response, UploadFile
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Font
from openpyxl.utils.exceptions import InvalidFileException
from pydantic import ValidationError
from sqlalchemy.orm import Session

logger = logging.getLogger(__name__)


def parse_user_import_form_options(form) -> dict[str, Any]:
    """Normalize shared password options from administrator import forms."""

    default_password = str(form.get("default_password") or "")
    force_password_change_raw = (
        str(form.get("force_password_change") or "true").strip().lower()
    )
    return {
        "default_password": default_password,
        "force_password_change": force_password_change_raw
        not in {"0", "false", "no", "n"},
    }


# Headers shared by the XLSX and CSV bulk-user templates.
BULK_USER_TEMPLATE_HEADERS: tuple[str, ...] = (
    "email",
    "first_name",
    "last_name",
)


REQUIRED_HEADERS = {"email", "first_name", "last_name"}
SUPPORTED_HEADERS = REQUIRED_HEADERS
MAX_BULK_USER_IMPORT_ROWS = 10_000
BulkUserRow = tuple[int, dict[str, Any]]


# Exception classes
class BulkUserTemplateError(ValueError):
    """Raised when a CSV or XLSX file cannot be processed for bulk import."""


# Helper functions
def _normalize_header(value: Any) -> str:
    """Normalize a header value to a string, stripping whitespace."""
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _row_has_values(row: Iterable[Any]) -> bool:
    """Check if a row contains any non-empty values."""
    return any(cell is not None and str(cell).strip() for cell in row)


def _canonicalize_header(value: str) -> str:
    """Canonicalize a header by lowercasing and replacing spaces with underscores."""
    return value.lower().replace(" ", "_")


def _validate_headers(headers: list[str], file_type: str) -> None:
    """Reject ambiguous or incomplete import headers."""
    populated_headers = [header for header in headers if header]
    duplicates = sorted(
        header
        for header, occurrence_count in Counter(populated_headers).items()
        if occurrence_count > 1
    )
    if duplicates:
        raise BulkUserTemplateError(
            f"{file_type} contains duplicate headers: {', '.join(duplicates)}"
        )

    unsupported_headers = sorted(set(populated_headers).difference(SUPPORTED_HEADERS))
    if unsupported_headers:
        raise BulkUserTemplateError(
            f"{file_type} contains unsupported headers: {', '.join(unsupported_headers)}"
        )

    missing_headers = REQUIRED_HEADERS.difference(populated_headers)
    if missing_headers:
        missing = ", ".join(sorted(missing_headers))
        raise BulkUserTemplateError(
            f"{file_type} must include headers: {', '.join(sorted(REQUIRED_HEADERS))}. "
            f"Missing: {missing}"
        )


def _build_row_data(
    headers: Iterable[str],
    row: Iterable[Any],
    *,
    file_type: str,
    row_number: int,
) -> dict[str, Any]:
    """Validate every populated cell before mapping a parsed import row."""
    header_values = tuple(headers)
    row_values = tuple(row)
    invalid_columns = [
        index + 1
        for index, value in enumerate(row_values)
        if value is not None
        and str(value).strip()
        and (
            index >= len(header_values)
            or not header_values[index]
            or header_values[index] not in SUPPORTED_HEADERS
        )
    ]
    if invalid_columns:
        raise BulkUserTemplateError(
            f"{file_type} row {row_number} contains data without a supported header "
            f"in columns: {', '.join(map(str, invalid_columns))}"
        )

    return {
        header: value.strip() if isinstance(value, str) else value
        for header, value in zip(header_values, row_values)
        if header
    }


def _bulk_import_error(message: str) -> dict[str, Any]:
    """Build the stable error result returned when an import file cannot be parsed."""
    return {
        "status": "error",
        "created_users": [],
        "errors": [message],
        "total_created": 0,
        "total_errors": 1,
    }


def _normalize_bulk_default_password(
    default_password: str | None,
    db: Session,
) -> str:
    """Normalize and validate the required seed for generated temporary passwords."""
    normalized_password = (default_password or "").strip()
    if not normalized_password:
        raise HTTPException(status_code=400, detail="Default password is required")
    _assert_password_policy(normalized_password, db)
    return normalized_password


def _build_bulk_temporary_password(
    default_password: str,
    requirements: dict[str, object],
) -> str:
    """Build a unique temporary password that satisfies the configured policy."""
    nonce = secrets.token_urlsafe(18)
    digest = hashlib.sha256(f"{default_password}:{nonce}".encode("utf-8")).hexdigest()

    min_len = max(1, int(requirements["min_len"]))
    min_special = max(0, int(requirements["min_special"]))
    min_upper = max(0, int(requirements["min_upper"]))
    min_lower = max(0, int(requirements["min_lower"]))
    min_num = max(0, int(requirements["min_num"]))
    special_characters = str(
        requirements.get("special_characters") or string.punctuation
    )
    safe_specials = "".join(
        char for char in "!#$%&()*+,-.:;<=>?@[]^_{|}~" if char in special_characters
    )
    if not safe_specials:
        safe_specials = special_characters[:1] or "!"

    # Seed every password with all common character classes, then add enough
    # characters to satisfy stricter administrator-defined requirements.
    chars = list(f"T{digest[:16]}{nonce}a1")
    chars.extend(secrets.choice(string.ascii_uppercase) for _ in range(min_upper))
    chars.extend(secrets.choice(string.ascii_lowercase) for _ in range(min_lower))
    chars.extend(secrets.choice(string.digits) for _ in range(min_num))
    chars.extend(secrets.choice(safe_specials) for _ in range(min_special))

    filler_chars = string.ascii_letters + string.digits + safe_specials
    while len(chars) < min_len:
        chars.append(secrets.choice(filler_chars))

    secrets.SystemRandom().shuffle(chars)
    return "".join(chars)


def _get_bulk_row_value(row_data: dict[str, Any], key: str) -> str:
    """Return a normalized string value from a parsed bulk-import row."""
    value = row_data.get(key)
    return (str(value) if value is not None else "").strip()


def _create_users_from_rows(
    rows: Iterable[BulkUserRow],
    db: Session,
    *,
    default_password: str | None = None,
    force_password_change: bool = True,
) -> dict[str, Any]:
    """Create users from normalized rows, committing each complete account once."""
    created_users: list[dict[str, str]] = []
    errors: list[str] = []

    validated_default_password = _normalize_bulk_default_password(default_password, db)
    temporary_password_requirements = get_password_policy_requirements(db)

    default_group_id = get_value_by_page_and_key(
        "login_general", "default_user_group", db
    )
    if not default_group_id:
        raise HTTPException(
            status_code=400,
            detail="Default user group is not configured. Please set it in settings.",
        )

    for row_number, row_data in rows:
        try:
            email = _get_bulk_row_value(row_data, "email").lower()
            generated_temporary_password = _build_bulk_temporary_password(
                validated_default_password,
                temporary_password_requirements,
            )
            first_name = _get_bulk_row_value(row_data, "first_name")
            last_name = _get_bulk_row_value(row_data, "last_name")

            if not email or not first_name or not last_name:
                errors.append(
                    f"Row {row_number}: Missing required fields (email, first_name, last_name)"
                )
                continue

            try:
                data = UserCreate(
                    email=email,
                    password=generated_temporary_password,
                    first_name=first_name,
                    last_name=last_name,
                )
            except ValidationError:
                errors.append(f"Row {row_number}: Invalid user data")
                continue

            if user_exists_by_email(db, data.email):
                errors.append(
                    f"Row {row_number}: User with email {data.email} already exists"
                )
                continue

            hashed_password = hash_password(data.password)
            user = create_user(
                db,
                data.email,
                hashed_password,
                data.first_name,
                data.last_name,
                USER_ROLE,
                default_group_id,
                commit=False,
            )

            # Password behavior comes only from the import options used by the
            # admin flow. Spreadsheet rows cannot alter account security state.
            must_change = bool(force_password_change)
            update_user_settings_bulk(
                user.id,
                {"security": {"has_to_change_password": must_change}},
                db,
                commit=False,
            )

            # Capture the response while the row is still in its transaction.
            # create_user() has already flushed and refreshed the account, so a
            # second refresh after commit would add no data and could falsely
            # report failure after the account had been permanently persisted.
            created_user = {
                "id": user.id,
                "email": user.email,
                "first_name": user.first_name,
                "last_name": user.last_name,
                "role": user.role,
            }
            created_user["temporary_password"] = generated_temporary_password

            # This is the row's final fallible database operation. Any failure
            # here can still be rolled back without leaving an unreported user.
            db.commit()
            created_users.append(created_user)
        except Exception as exc:
            try:
                db.rollback()
            except Exception as rollback_exc:
                logger.error(
                    "Failed to roll back bulk-user row %s (error_type=%s)",
                    row_number,
                    type(rollback_exc).__name__,
                )
            # SQL exceptions can contain bound account data, so log only the
            # exception type and keep details out of client output.
            logger.error(
                "Failed to create bulk-user row %s (error_type=%s)",
                row_number,
                type(exc).__name__,
            )
            errors.append(f"Row {row_number}: Unable to create user")

    return {
        "status": "success",
        "created_users": created_users,
        "errors": errors,
        "total_created": len(created_users),
        "total_errors": len(errors),
    }


# -------------------
# XLSX bulk import
# -------------------
def generate_xlsx_template() -> bytes:
    """Return an XLSX workbook (as bytes) containing the bulk user headers."""

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Users"

    header_font = Font(bold=True)
    worksheet.append(list(BULK_USER_TEMPLATE_HEADERS))

    for cell in worksheet[1]:  # type: ignore[index]
        if isinstance(cell, Cell):
            cell.font = header_font

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _template_response(
    content: bytes,
    filename: str,
    media_type: str,
    extra_headers: dict[str, str] | None = None,
) -> Response:
    """Build a cacheable template download response."""
    now = datetime.now(timezone.utc)
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"',
        "Cache-Control": "public, max-age=86400",
        "ETag": hashlib.md5(content, usedforsecurity=False).hexdigest(),
        "Last-Modified": format_datetime(now),
        "Expires": format_datetime(now + timedelta(days=1)),
    }
    headers.update(extra_headers or {})
    return Response(content=content, media_type=media_type, headers=headers)


def admin_get_users_xlsx_template(_locale: str):
    """Return the locale-independent XLSX template."""
    return _template_response(
        generate_xlsx_template(),
        "users_template.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def iter_xlsx_rows(xlsx_bytes: bytes) -> Iterator[BulkUserRow]:
    """Yield (row_number, row_dict) pairs from the first worksheet of the workbook."""

    try:
        workbook = load_workbook(BytesIO(xlsx_bytes), data_only=True, read_only=True)
    except InvalidFileException as exc:
        raise BulkUserTemplateError(
            "Uploaded file is not a valid XLSX workbook"
        ) from exc
    except Exception as exc:  # pragma: no cover - unexpected load errors
        raise BulkUserTemplateError("Unable to read XLSX workbook") from exc

    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)

    header_row: tuple[Any, ...] | None = None
    header_row_index: int | None = None
    for idx, row in enumerate(rows, start=worksheet.min_row):
        if _row_has_values(row):
            header_row = row
            header_row_index = idx
            break

    if header_row is None or header_row_index is None:
        raise BulkUserTemplateError("XLSX workbook is empty")

    headers = [_canonicalize_header(_normalize_header(value)) for value in header_row]
    _validate_headers(headers, "XLSX")

    for index, row in enumerate(rows, start=header_row_index + 1):
        if not _row_has_values(row):
            # Skip entirely empty rows but keep row numbering consistent
            continue

        yield (
            index,
            _build_row_data(
                headers,
                row,
                file_type="XLSX",
                row_number=index,
            ),
        )


SEMICOLON_LANGUAGES = frozenset(
    {"de", "fr", "it", "es", "pt", "nl", "pl", "ru", "ar", "tr"}
)
DEFAULT_DELIMITER = ","


def _normalize_locale(value: str) -> str:
    """Normalizes a locale string to ll_CC format, e.g., 'en-us' -> 'en_US'."""
    normalized = value.strip().replace("-", "_")
    if not re.fullmatch(r"[a-zA-Z0-9_]*", normalized):
        raise HTTPException(
            status_code=400, detail="Locale contains invalid characters"
        )

    if "_" in normalized:
        lang, _, region = normalized.partition("_")
        if lang and region:
            return f"{lang.lower()}_{region.upper()}"

    return normalized.lower()


def _get_delimiter_for_locale(locale: str) -> str:
    """Get the CSV delimiter for a given locale, with language fallback."""
    language = _normalize_locale(locale).split("_", 1)[0]
    return ";" if language in SEMICOLON_LANGUAGES else DEFAULT_DELIMITER


def _parse_accept_language(accept_language: str | None) -> str:
    """Parse Accept-Language header and return the best locale."""
    if not accept_language:
        return "en_US"

    languages: list[tuple[str, float]] = []
    for part in accept_language.split(","):
        part = part.strip()
        if not part:
            continue

        if ";" in part:
            lang, _, q_part = part.partition(";")
            lang = lang.strip()
            q_match = re.search(r"q=([0-9.]+)", q_part)
            q = float(q_match.group(1)) if q_match else 1.0
        else:
            lang = part
            q = 1.0

        # Wildcards are valid language ranges, but contain no delimiter signal.
        if lang and lang != "*":
            languages.append((lang, q))

    languages.sort(key=lambda x: x[1], reverse=True)

    for lang, _ in languages:
        try:
            return _normalize_locale(lang)
        except HTTPException:
            # A malformed client preference must not break template download.
            continue

    return "en_US"


# -------------------
# CSV Template Generation
# -------------------
def generate_csv_template(delimiter: str = ",") -> bytes:
    """Return a CSV template with the bulk user headers."""
    output = StringIO()
    writer = csv.writer(output, delimiter=delimiter, quoting=csv.QUOTE_MINIMAL)
    writer.writerow(BULK_USER_TEMPLATE_HEADERS)
    return output.getvalue().encode("utf-8-sig")


def admin_get_users_csv_template(accept_language: str | None = None):
    """Generate and return a CSV template with locale-appropriate delimiter."""
    locale = _parse_accept_language(accept_language)
    delimiter = _get_delimiter_for_locale(locale)
    content = generate_csv_template(delimiter)
    delimiter_name = "semicolon" if delimiter == ";" else "comma"
    return _template_response(
        content,
        f"users_template_{delimiter_name}.csv",
        "text/csv; charset=utf-8",
        {"X-CSV-Delimiter": delimiter, "X-Detected-Locale": locale},
    )


# -------------------
# CSV Parsing with Auto-Detection
# -------------------
def _detect_csv_delimiter(sample: str) -> str:
    """Detect the delimiter used in a CSV sample using csv.Sniffer."""
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        return dialect.delimiter
    except csv.Error:
        comma_count = sample.count(",")
        semicolon_count = sample.count(";")
        return ";" if semicolon_count > comma_count else ","


def iter_csv_rows(csv_bytes: bytes) -> Iterator[BulkUserRow]:
    """Yield (row_number, row_dict) pairs from CSV content with auto-delimiter detection."""
    try:
        # CP1252 must precede Latin-1 because Latin-1 accepts every byte and
        # would otherwise decode Windows punctuation as control characters.
        for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
            try:
                content = csv_bytes.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            raise BulkUserTemplateError(
                "Unable to decode CSV file. Please use UTF-8 encoding."
            )

        lines = content.strip().split("\n")
        if not lines:
            raise BulkUserTemplateError("CSV file is empty")

        sample = "\n".join(lines[: min(5, len(lines))])
        delimiter = _detect_csv_delimiter(sample)

        reader = csv.reader(StringIO(content), delimiter=delimiter)

        header_row: list[str] = []
        header_row_index = 0
        for idx, row in enumerate(reader, start=1):
            if any(cell.strip() for cell in row):
                header_row = row
                header_row_index = idx
                break

        if not header_row:
            raise BulkUserTemplateError("CSV file is empty or has no header row")

        headers = [_canonicalize_header(h.strip()) for h in header_row]
        _validate_headers(headers, "CSV")

        for idx, row in enumerate(reader, start=header_row_index + 1):
            if not any(cell.strip() for cell in row):
                continue

            yield (
                idx,
                _build_row_data(
                    headers,
                    row,
                    file_type="CSV",
                    row_number=idx,
                ),
            )

    except BulkUserTemplateError:
        raise
    except Exception as exc:
        logger.exception("Unable to parse bulk-user CSV")
        raise BulkUserTemplateError("Unable to read CSV file") from exc


def _create_users_from_file(
    contents: bytes,
    db: Session,
    parser: Callable[[bytes], Iterator[BulkUserRow]],
    *,
    default_password: str | None = None,
    force_password_change: bool = True,
) -> dict[str, Any]:
    """Parse one supported file format and invoke the shared row importer."""
    try:
        rows: list[BulkUserRow] = []
        for parsed_row in parser(contents):
            if len(rows) >= MAX_BULK_USER_IMPORT_ROWS:
                raise BulkUserTemplateError(
                    f"Import file exceeds the {MAX_BULK_USER_IMPORT_ROWS:,} user limit"
                )
            rows.append(parsed_row)
    except BulkUserTemplateError as exc:
        return _bulk_import_error(str(exc))
    except Exception:
        logger.exception("Unable to parse bulk-user upload")
        return _bulk_import_error("Unable to read import file")

    return _create_users_from_rows(
        rows,
        db,
        default_password=default_password,
        force_password_change=force_password_change,
    )


def create_users_from_xlsx(
    contents: bytes,
    db: Session,
    default_password: str | None = None,
    force_password_change: bool = True,
) -> dict[str, Any]:
    """Create users from XLSX content."""
    return _create_users_from_file(
        contents,
        db,
        iter_xlsx_rows,
        default_password=default_password,
        force_password_change=force_password_change,
    )


def create_users_from_csv(
    contents: bytes,
    db: Session,
    *,
    default_password: str | None = None,
    force_password_change: bool = True,
) -> dict[str, Any]:
    """Create users from CSV content."""
    return _create_users_from_file(
        contents,
        db,
        iter_csv_rows,
        default_password=default_password,
        force_password_change=force_password_change,
    )


MAX_UPLOAD_BYTES = 5 * 1024 * 1024


def admin_upload_users_file(
    file: UploadFile,
    db: Session,
    *,
    default_password: str | None = None,
    force_password_change: bool = True,
):
    """Validate, read, and import either supported bulk-user file format."""
    try:
        if not file.filename:
            raise HTTPException(status_code=400, detail="No file provided")

        normalized_password = (default_password or "").strip()
        if not normalized_password:
            raise HTTPException(status_code=400, detail="Default password is required")

        filename = file.filename.lower()
        if filename.endswith(".xlsx"):
            importer, file_type = create_users_from_xlsx, "xlsx"
        elif filename.endswith(".csv"):
            importer, file_type = create_users_from_csv, "csv"
        else:
            raise HTTPException(
                status_code=400,
                detail="Unsupported file format. Please upload an XLSX or CSV file.",
            )

        contents = file.file.read(MAX_UPLOAD_BYTES + 1)
        if len(contents) > MAX_UPLOAD_BYTES:
            raise HTTPException(status_code=413, detail="File exceeds the 5 MB limit")

        result = importer(
            contents,
            db,
            default_password=normalized_password,
            force_password_change=force_password_change,
        )
        return result, file_type
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Unable to process bulk-user upload")
        raise HTTPException(
            status_code=500,
            detail="An unexpected error occurred while processing the file",
        ) from exc
    finally:
        file.file.close()
